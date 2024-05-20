import os
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv
from io import TextIOWrapper
from conftest import CURRENT_DIRECTORY

RESOURCES_DIR = os.path.join(CURRENT_DIRECTORY, "resources")


def test_archive_pdf():
    with ZipFile(os.path.join(RESOURCES_DIR, "files.zip"), "r") as zip_file:
        # print(zip_file.namelist())
        reader = PdfReader(zip_file.open("test.pdf"))
        text = reader.pages[1].extract_text()
        assert "Работа подготовлена на кафедре прикладной математики" in text


def test_archive_xlsx():
    with ZipFile(os.path.join(RESOURCES_DIR, "files.zip"), "r") as zip_file:
        workbook = load_workbook(zip_file.open("test.xlsx"))
        sheet = workbook.active
        text = sheet.cell(row=15, column=3).value
        assert "Значения искомых перемен." in text


def test_archive_csv():
    with ZipFile(os.path.join(RESOURCES_DIR, "files.zip"), "r") as zip_file:
        with zip_file.open("test.csv") as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, "Windows-1251"), delimiter=';'))
            # print(list(csvreader))
            second_row = csvreader[1]
            # print(second_row)

            assert "Произв. оборуд. в нормочасах" in second_row[0]
            assert second_row[3] == "1"
